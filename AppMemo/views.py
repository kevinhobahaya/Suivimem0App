from django.shortcuts import render, redirect
from django.conf import settings
from django.core.mail import send_mail
from .models import *
from decimal import Decimal
from django.shortcuts import render, redirect
from .models import Concerne
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from .models import EtatSortie
from django.template.loader import render_to_string
import weasyprint
from django.conf import settings
import os
from django.http import HttpResponse, HttpResponseNotFound
from django.contrib.auth.models import User,auth 
from django.shortcuts import redirect
from .forms import EtatSortieForm
# Create your views here.
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import EtatSortie
from django.contrib.auth import authenticate, login
from django.db.models import Sum
from django.shortcuts import render
from .models import Concerne
import json
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from django.db.models import Sum, F, FloatField
from .models import Concerne
import os
from datetime import datetime
from django.db.models import Sum, F, FloatField
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Concerne
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from .models import Concerne
from django.db.models import Sum, F, FloatField

def dashboard_concerne(request):
    # 1️⃣ Total brut par mois
    total_par_mois = (
        Concerne.objects
        .values('mois_ref')
        .annotate(total=Sum('total_brut'))
        .order_by('mois_ref')
    )
    mois = [item['mois_ref'] for item in total_par_mois]
    total_brut = [float(item['total']) for item in total_par_mois]

    # 2️⃣ Total par importateur
    total_par_importateur = (
        Concerne.objects
        .values('importateur')
        .annotate(total=Sum('total_brut'))
        .order_by('-total')[:10]  # top 10
    )
    importateurs = [item['importateur'] for item in total_par_importateur]
    total_importateur = [float(item['total']) for item in total_par_importateur]

    # 3️⃣ Frais par mois
    frais_par_mois = (
        Concerne.objects
        .values('mois_ref')
        .annotate(
            controle=Sum('frais_controle'),
            labo=Sum('frais_analyse_labo'),
            tva=Sum('tva')
        )
        .order_by('mois_ref')
    )
    mois_frais = [item['mois_ref'] for item in frais_par_mois]
    frais_controle = [float(item['controle']) for item in frais_par_mois]
    frais_labo = [float(item['labo']) for item in frais_par_mois]
    tva = [float(item['tva']) for item in frais_par_mois]

    context = {
        'mois': json.dumps(mois),
        'total_brut': json.dumps(total_brut),
        'importateurs': json.dumps(importateurs),
        'total_importateur': json.dumps(total_importateur),
        'mois_frais': json.dumps(mois_frais),
        'frais_controle': json.dumps(frais_controle),
        'frais_labo': json.dumps(frais_labo),
        'tva': json.dumps(tva),
    }

    return render(request, 'admin/index.html', context)






def createUser(request):
    user = User.objects.all()
    if request.method=='POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm = request.POST.get('confirm')
        role_name = request.POST.get('role')  # Le nom du rôle à associer

        if password == confirm:
            if User.objects.filter(email=email).exists():
                return HttpResponseNotFound("L'utilisateur existe déjà")
            else:
                # Créez l'utilisateur
                user = User(
                    email=email,
                    username=username,
                    is_superuser=1,
                    is_staff=0  # Vous pouvez ajuster cette valeur si nécessaire
                )
                user.set_password(password)
                user.save()
                # Envoyez un e-mail de bienvenue
                # subject = 'Bienvenu à GestionTalent'
                # message = f"Bonjour {username}, Merci pour votre souscription. \n Veuillez cliquer sur ce lien pour configurer votre compte \n http://127.0.0.1:8000/login"
                # email_from = settings.EMAIL_HOST_USER
                # recipient_list = [email]
                # send_mail(subject, message, email_from, recipient_list)
                # Récupérez ou créez le rôle
                profil, created = Profil.objects.get_or_create(name=role_name)
                
                # Associez le rôle à l'utilisateur
                user.desc.add(profil)
                
                return redirect('AppMemo:login')
        else:
            return HttpResponseNotFound("Les mots de passe ne correspondent pas")

    return HttpResponse("Méthode non autorisée")
# login

def loginUser(request):
    if request.method=='POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = auth.authenticate(username = username, password = password)

        if user is not None:
            login(request, user)
            
            # Débogage : vérifier les rôles
            print(f'Roles of user {user.username}: {user.roles.values_list("name", flat=True)}')
            
            # Vérifiez les rôles de l'utilisateur
            if user.desc.filter(name='Admin').exists():
                return redirect('AppMemo:home')
            elif user.star.filter(name='Sec').exists():
                return redirect('AppMemo:detail')
            elif user.roles.filter(name='Agent').exists():
                return redirect('AppMemo:dashboard_agent')
            else:
                return redirect('AppMemo:Accueil_admin')  # Vous pouvez ajouter d'autres rôles ici

        return HttpResponse("Identifiants invalides")
    


def etat_sortie_list(request):
    search_query = request.GET.get('q', '')
    etats = EtatSortie.objects.all()

    if search_query:
        etats = etats.filter(num_e__icontains=search_query)  # __icontains pour recherche partielle

    paginator = Paginator(etats, 10)  # 10 éléments par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    detail = Concerne.objects.all()

    context = {
        'detail': page_obj.object_list,
        'page_obj': page_obj,
        'search_query': search_query,
        'detail': detail
    }
    return render(request, 'controle/detailEtat.html', context)

def dashboardPage(request):
    concernes = Concerne.objects.all()
    # Extraire les données
    num_ref = [concerne.num_ref for concerne in concernes]
    mois_ref = [concerne.mois_ref for concerne in concernes]

    context = {
        'num_ref': num_ref,
        'mois_ref': mois_ref
    }
    return render(request,'admin/index.html', context)

def concernePage(request):
    return render(request, 'admin/concerne.html')

def pageDetail(request):
    detail = Concerne.objects.all()
    return render(request, 'controle/detailEtat.html',{'detail': detail})

def CreateConcerne(request):
    if request.method == 'POST':
        importateur = request.POST.get('importateur')
        container = request.POST.get('container')
        marchandise = request.POST.get('marchandise')
        quantite = request.POST.get('quantite')
        poids = request.POST.get('poids')
        fob = request.POST.get('fob')
        cif = request.POST.get('cif')
        num_bl = request.POST.get('num_bl')
        num_fbr_aa = request.POST.get('num_fbr_aa')
        plaque = request.POST.get('plaque')
        num_e = request.POST.get('num_e')
        origine = request.POST.get('origine')
        provenance = request.POST.get('provenance')
        mois_ref= request.POST.get('mois_ref')
        num_ref = request.POST.get('num_ref')
        transitaire = request.POST.get('transitaire')
        destination = request.POST.get('destination')
        frais_controle = Decimal(request.POST.get('frais_controle', 0))
        frais_analyse_labo = Decimal(request.POST.get('frais_analyse_labo', 0))
        tva = Decimal(request.POST.get('tva', 0))
        nb_no = request.POST.get('nb_no')
        recu_no = request.POST.get('recu_no')
        bv_num = request.POST.get('bv_num')
        liquidation_no = request.POST.get('liquidation_no')
        travail_laboratoire = request.POST.get('travail_laboratoire')
        num_av = request.POST.get('num_av')
       
        # Calcul du total_brut
        total_brut = frais_controle + frais_analyse_labo + tva
        form = Concerne(
            importateur = importateur,
            container = container,
            marchandise = marchandise,
            quantite = quantite,
            poids = poids,
            fob = fob,
            cif = cif,
            num_bl = num_bl,
            num_fbr_aa = num_fbr_aa,
            plaque = plaque,
            num_e = num_e,
            origine = origine,
            provenance = provenance,
            num_ref= num_ref,
            mois_ref = mois_ref,
            transitaire = transitaire,
            destination = destination,
            frais_controle=frais_controle,
            frais_analyse_labo=frais_analyse_labo,
            tva=tva,
            total_brut=total_brut,
            nb_no=nb_no,
            recu_no=recu_no,
            bv_num = bv_num,
            liquidation_no=liquidation_no,
            travail_laboratoire=travail_laboratoire,
            num_av = num_av,
        )
        form.save()
        return redirect('AppMemo:detail')


def UpdateConcerne(request,pk):
    if request.method == 'POST':
        importateur = request.POST.get('importateur')
        container = request.POST.get('container')
        marchandise = request.POST.get('marchandise')
        quantite = request.POST.get('quantite')
        poids = request.POST.get('poids')
        fob = request.POST.get('fob')
        cif = request.POST.get('cif')
        num_bl = request.POST.get('num_bl')
        num_fbr_aa = request.POST.get('num_fbr_aa')
        plaque = request.POST.get('plaque')
        num_e = request.POST.get('num_e')
        origine = request.POST.get('origine')
        provenance = request.POST.get('provenance')
        mois_ref= request.POST.get('mois_ref')
        num_ref = request.POST.get('num_ref')
        transitaire = request.POST.get('transitaire')
        destination = request.POST.get('destination')
        frais_controle = Decimal(request.POST.get('frais_controle', 0))
        frais_analyse_labo = Decimal(request.POST.get('frais_analyse_labo', 0))
        tva = Decimal(request.POST.get('tva', 0))
        nb_no = request.POST.get('nb_no')
        recu_no = request.POST.get('recu_no')
        bv_num = request.POST.get('bv_num')
        liquidation_no = request.POST.get('liquidation_no')
        travail_laboratoire = request.POST.get('travail_laboratoire')
        num_av = request.POST.get('num_av')
       
        # Calcul du total_brut
        total_brut = frais_controle + frais_analyse_labo + tva
        form = Concerne(
            id = pk,
            importateur = importateur,
            container = container,
            marchandise = marchandise,
            quantite = quantite,
            poids = poids,
            fob = fob,
            cif = cif,
            num_bl = num_bl,
            num_fbr_aa = num_fbr_aa,
            plaque = plaque,
            num_e = num_e,
            origine = origine,
            provenance = provenance,
            num_ref= num_ref,
            mois_ref = mois_ref,
            transitaire = transitaire,
            destination = destination,
            frais_controle=frais_controle,
            frais_analyse_labo=frais_analyse_labo,
            tva=tva,
            total_brut=total_brut,
            nb_no=nb_no,
            recu_no=recu_no,
            bv_num = bv_num,
            liquidation_no=liquidation_no,
            travail_laboratoire=travail_laboratoire,
            num_av = num_av,
        )
        form.save()
        return redirect('AppMemo:detail')

def updateConcerneDetail(request,pk):
    detail = Concerne.objects.get(id=pk)
    context = {
        'detail':detail
    }
    return render(request,'controle/detailEtat.html',context)            

# etat de sortie

def etat_sortie_detail(request, pk):
    etat = get_object_or_404(EtatSortie, pk=pk)
    return render(request, 'controle/etatsortie.html', {'etat': etat})

def etat_sortie_pdf(request, pk):
    etat = get_object_or_404(EtatSortie, pk=pk)
    html_string = render_to_string('controle/etatsortie.html', {'etat': etat})
    # CSS (fichier statique)
    css_path = os.path.join(settings.BASE_DIR, 'controle', 'static', 'controle', 'css', 'print.css')
    pdf = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(stylesheets=[weasyprint.CSS(css_path)])
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="etat_sortie_{etat.id}.pdf"'
    return response




def etat_create(request):
    if request.method == 'POST':
        form = EtatSortieForm(request.POST)
        if form.is_valid():
            etat = form.save()
            return redirect('controle:etat_detail', pk=etat.pk)
    else:
        form = EtatSortieForm()
    return render(request, 'controle/etat_form.html', {'form': form})


def loginPage(request):
    return render(request, 'admin/login.html')

def comptePage(request):
    return render(request, 'admin/compte.html')



# views.py
from django.shortcuts import render, get_object_or_404
from .models import Concerne
from django.shortcuts import get_object_or_404, render

def memo_interne_view(request, uuid):
    rapport = get_object_or_404(Concerne, uuid=uuid)
    return render(request, 'controle/memo_interne.html', {'rapport': rapport})


# exportation en excel



def export_stat_mensuel_pdf(request):
    # Récupération des statistiques comme pour Excel
    stats = (
        Concerne.objects
        .values('mois_ref')
        .annotate(
            total_brut=Sum('total_brut'),
            frais_controle=Sum('frais_controle'),
            frais_labo=Sum('frais_analyse_labo'),
            tva=Sum('tva'),
            total_general=Sum(
                F('total_brut') + F('frais_controle') + F('frais_analyse_labo') + F('tva'),
                output_field=FloatField()
            )
        )
        .order_by('mois_ref')
    )

    # Totaux généraux
    total_global = {
        'total_brut': sum(item['total_brut'] for item in stats),
        'frais_controle': sum(item['frais_controle'] for item in stats),
        'frais_labo': sum(item['frais_labo'] for item in stats),
        'tva': sum(item['tva'] for item in stats),
        'total_general': sum(item['total_general'] for item in stats),
    }

    # Configuration de la réponse HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="statistiques_mensuelles.pdf"'

    # Création du document PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Titre
    elements.append(Paragraph("Statistiques Mensuelles des Importations", styles['Title']))
    elements.append(Spacer(1, 12))

    # En-têtes du tableau
    data = [["Mois", "Total Brut", "Frais Contrôle", "Frais Labo", "TVA", "Total Général"]]

    # Contenu du tableau
    for item in stats:
        data.append([
            item['mois_ref'],
            f"{item['total_brut']:.2f}",
            f"{item['frais_controle']:.2f}",
            f"{item['frais_labo']:.2f}",
            f"{item['tva']:.2f}",
            f"{item['total_general']:.2f}",
        ])

    # Ligne finale - Total général
    data.append(["", "", "", "", "", ""])
    data.append([
        "TOTAL",
        f"{total_global['total_brut']:.2f}",
        f"{total_global['frais_controle']:.2f}",
        f"{total_global['frais_labo']:.2f}",
        f"{total_global['tva']:.2f}",
        f"{total_global['total_general']:.2f}",
    ])

    # Création du tableau
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#343a40")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    doc.build(elements)

    return response

from django.db.models import Sum, F, FloatField
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Concerne

def tableau_stat_mensuel(request):
    # Annotation pour chaque ligne
    stats_query = (
        Concerne.objects.values('importateur', 'mois_ref')
        .annotate(
            total_brut=Sum('total_brut'),
            frais_controle=Sum('frais_controle'),
            frais_analyse_labo=Sum('frais_analyse_labo'),
            tva=Sum('tva')
        )
        .order_by('mois_ref')
    )

    # Pagination
    paginator = Paginator(stats_query, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calcul des totaux globaux en Python
    total_global = {
        'total_brut': sum(item['total_brut'] or 0 for item in stats_query),
        'frais_controle': sum(item['frais_controle'] or 0 for item in stats_query),
        'frais_labo': sum(item['frais_analyse_labo'] or 0 for item in stats_query),
        'tva': sum(item['tva'] or 0 for item in stats_query),
    }
    total_global['total_general'] = (
        total_global['total_brut'] +
        total_global['frais_controle'] +
        total_global['frais_labo'] +
        total_global['tva']
    )

    return render(request, 'admin/tableaustatmensuel.html', {
        'stats': page_obj,
        'total_global': total_global
    })


# Export PDF paysage
from django.http import HttpResponse
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from django.db.models import F, Sum, ExpressionWrapper, FloatField
from .models import Concerne

def export_stat_mensuel_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="statistiques_mensuelles.pdf"'

    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)

    # -----------------------------
    # En-tête avec logo
    # -----------------------------
    logo_path = 'static/images/logo_occ.png'  # chemin vers le logo OCC
    try:
        logo = ImageReader(logo_path)
        c.drawImage(logo, 50, height - 80, width=60, height=60, preserveAspectRatio=True)
    except:
        pass  # si le logo n'existe pas, on continue

    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, height - 50, "REPUBLIQUE DEMOCRATIQUE DU CONGO")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, height - 70, "MINISTERE DU COMMERCE")
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(width / 2, height - 90, "STATISTIQUES MENSUELLES DES IMPORTATIONS")

    # -----------------------------
    # Tableau
    # -----------------------------
    y = height - 120
    headers = ['Mois', 'Total Brut', 'Frais Contrôle', 'Frais Analyse', 'TVA', 'Total']
    x_positions = [50, 150, 250, 350, 450, 550]

    c.setFont("Helvetica-Bold", 10)
    for x, header in zip(x_positions, headers):
        c.drawString(x, y, header)
    y -= 20

    total_expr = ExpressionWrapper(
        F('total_brut') + F('frais_controle') + F('frais_analyse_labo') + F('tva'),
        output_field=FloatField()
    )

    stat_mensuel = Concerne.objects.values('mois_ref').annotate(
        total_sum=Sum(total_expr),
        total_brut_sum=Sum('total_brut'),
        frais_controle_sum=Sum('frais_controle'),
        frais_analyse_labo_sum=Sum('frais_analyse_labo'),
        tva_sum=Sum('tva')
    ).order_by('mois_ref')

    c.setFont("Helvetica", 10)
    for stat in stat_mensuel:
        c.drawString(x_positions[0], y, str(stat['mois_ref']))
        c.drawString(x_positions[1], y, f"{stat['total_brut_sum']:.2f}")
        c.drawString(x_positions[2], y, f"{stat['frais_controle_sum']:.2f}")
        c.drawString(x_positions[3], y, f"{stat['frais_analyse_labo_sum']:.2f}")
        c.drawString(x_positions[4], y, f"{stat['tva_sum']:.2f}")
        c.drawString(x_positions[5], y, f"{stat['total_sum']:.2f}")
        y -= 20

        if y < 80:  # nouvelle page si on arrive en bas
            c.showPage()
            y = height - 50

    # -----------------------------
    # Bas de page : signatures
    # -----------------------------
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, 50, "Le Chef de détachement")
    c.drawString(width - 200, 50, "Le Chef de bureau")

    c.save()
    return response


# Export Excel
def export_stat_mensuel_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistiques Mensuelles"

    headers = ['Mois', 'Total Brut', 'Frais Contrôle', 'Frais Analyse', 'TVA', 'Total']
    ws.append(headers)

    total_expr = ExpressionWrapper(
        F('total_brut') + F('frais_controle') + F('frais_analyse_labo') + F('tva'),
        output_field=FloatField()
    )

    stat_mensuel = Concerne.objects.values('mois_ref').annotate(
        total_sum=Sum(total_expr),
        total_brut_sum=Sum('total_brut'),
        frais_controle_sum=Sum('frais_controle'),
        frais_analyse_labo_sum=Sum('frais_analyse_labo'),
        tva_sum=Sum('tva')
    ).order_by('mois_ref')

    for stat in stat_mensuel:
        ws.append([
            stat['mois_ref'],
            stat['total_brut_sum'],
            stat['frais_controle_sum'],
            stat['frais_analyse_labo_sum'],
            stat['tva_sum'],
            stat['total_sum']
        ])

    # Ajustement largeur colonnes pour éviter que tout soit collé
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=statistiques_mensuelles.xlsx'
    wb.save(response)
    return response



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import Concerne


def export_stat_mensuel_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="statistiques_mensuelles.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Statistiques Importations"

    # --- En-tête ---
    ws.merge_cells('A1:I1')
    ws['A1'] = "REPUBLIQUE DEMOCRATIQUE DU CONGO"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = "OFFICE CONGOLAIS DE CONTRÔLE"
    ws['A2'].font = Font(bold=True, size=13)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:I3')
    ws['A3'] = "DIRECTION PROVINCIALE DU NORD-KIVU"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:I4')
    ws['A4'] = "AGENCE DE BENI - DÉTACHEMENT DE KASINDI"
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A6:I6')
    ws['A6'] = "STATISTIQUES MENSUELLES DES MEMO INTERNES"
    ws['A6'].font = Font(bold=True, size=12)
    ws['A6'].alignment = Alignment(horizontal='center')

    # --- Titres du tableau ---
    headers = [
        "Importateur", "Plaque", "Date", "Marchandise",
        "Quantité", "Poids", "Num E", "Origine", "Provenance"
    ]
    start_row = 8

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # --- Données du tableau ---
    data = Concerne.objects.all().order_by('date')

    for row_num, item in enumerate(data, start=start_row + 1):
        ws.cell(row=row_num, column=1, value=str(item.importateur or ""))
        ws.cell(row=row_num, column=2, value=item.plaque or "")
        ws.cell(row=row_num, column=3, value=str(item.date or ""))
        ws.cell(row=row_num, column=4, value=item.marchandise or "")
        ws.cell(row=row_num, column=5, value=item.quantite or "")
        ws.cell(row=row_num, column=6, value=item.poids or "")
        ws.cell(row=row_num, column=7, value=item.num_e or "")
        ws.cell(row=row_num, column=8, value=item.origine or "")
        ws.cell(row=row_num, column=9, value=item.provenance or "")

    # --- Ajustement des colonnes ---
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    # --- Signatures ---
    last_row = start_row + len(data) + 3
    ws.merge_cells(f'A{last_row}:C{last_row}')
    ws[f'A{last_row}'] = "LE CHEF DE BUREAU"
    ws[f'A{last_row}'].font = Font(bold=True)
    ws[f'A{last_row}'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'G{last_row}:I{last_row}')
    ws[f'G{last_row}'] = "LE CHEF DE DÉTACHEMENT"
    ws[f'G{last_row}'].font = Font(bold=True)
    ws[f'G{last_row}'].alignment = Alignment(horizontal='center')

    wb.save(response)
    return response

# -----------------------------
# Export PDF
# -----------------------------
def export_stat_mensuel_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="statistiques_mensuelles.pdf"'

    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)

    # --------- En-tête avec logo et titre ---------
    logo_path = 'static/image/logo.png'
    try:
        logo = ImageReader(logo_path)
        c.drawImage(logo, 50, height - 80, width=60, height=60, preserveAspectRatio=True)
    except:
        pass

    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, height - 50, "REPUBLIQUE DEMOCRATIQUE DU CONGO")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, height - 70, "OFFICE CONGOLAIS DE CONTRÔLE")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, height - 90, "DIRECTION PROVINCIALE DU NORD-KIVU")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, height - 110, "AGENCE DE BENI")
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, height - 130, "DÉTACHEMENT DE KASINDI")
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(width / 2, height - 160, "STATISTIQUES MENSUELLES DES MEMO INTERNES")

    # --------- Tableau ---------
    y = height - 190
    headers = [
        "Importateur", "Plaque", "Date", "Marchandise",
        "Quantité", "Poids", "Num E", "Origine", "Provenance"
    ]
    x_positions = [50, 200, 280, 380, 460, 520, 590, 650, 740]

    c.setFont("Helvetica-Bold", 9)
    for x, header in zip(x_positions, headers):
        c.drawString(x, y, header)
    y -= 15
    c.setFont("Helvetica", 8)

    data = Concerne.objects.all().order_by('date')

    for item in data:
        if y < 60:
            c.showPage()
            c.setFont("Helvetica-Bold", 9)
            for x, header in zip(x_positions, headers):
                c.drawString(x, height - 40, header)
            c.setFont("Helvetica", 8)
            y = height - 60

        c.drawString(x_positions[0], y, str(item.importateur or ""))
        c.drawString(x_positions[1], y, str(item.plaque or ""))
        c.drawString(x_positions[2], y, str(item.date or ""))
        c.drawString(x_positions[3], y, str(item.marchandise or ""))
        c.drawString(x_positions[4], y, str(item.quantite or ""))
        c.drawString(x_positions[5], y, str(item.poids or ""))
        c.drawString(x_positions[6], y, str(item.num_e or ""))
        c.drawString(x_positions[7], y, str(item.origine or ""))
        c.drawString(x_positions[8], y, str(item.provenance or ""))
        y -= 15

    # --- Signatures en bas ---
    y = 50
    c.setFont("Helvetica-Bold", 10)
    c.drawString(80, y, "LE CHEF DE BUREAU")
    c.drawString(width - 250, y, "LE CHEF DE DÉTACHEMENT")

    c.save()
    return response